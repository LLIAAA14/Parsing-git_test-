import requests
from bs4 import BeautifulSoup
import openpyxl 

def get_html(url):
    house = requests.get(url)
    return house.text

def normal_str(s):
    r = ' '.join(s.split())
    return r

def write_exel(rooms_list):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    keys = rooms_list[0].keys()


    for col, key in enumerate(keys, start=1):
        sheet.cell(row=1, column=col, value=key)


    for row, room in enumerate(rooms_list, start=2):
        for col, key in enumerate(keys, start=1):
            sheet.cell(row=row, column=col, value=room[key])


    workbook.save("rooms_kaliningrad.xlsx")
    

def get_data(html):
    soup = BeautifulSoup(html, 'html.parser')
    rooms = soup.find_all('div', class_='card')
    rooms_list = []
    for room in rooms:
        try:
            type_room = room.find(class_='object-hotel__type').text
            name = room.find(class_='card-content__object-type js-name-click').text
            price = normal_str(room.find(class_='price js-popover popover-wrap').text)
            facilities = normal_str(room.find(class_='facilities__main popover-wrap popover-top').text)
            urls = room.find(class_='card-content').find('a').get('href')
            address = normal_str(room.find(class_='address__text has-tooltip').text)

        except Exception:
            type_room = 'Нет данных'
            name = 'Нет данных'
            facilities = 'Нет данных'
            address = 'Нет данных'
        
    
        rooms_dict = {'тип квартиры': type_room,
                      'название': name,
                      'цена': price,
                      'удобства': facilities,
                      'адрес': address,
                      'ссылка': urls}


        rooms_list.append(rooms_dict)
    write_exel(rooms_list)
    
              
def main():
    url = 'https://kaliningrad.sutochno.ru/'
    get_data(get_html(url))

if __name__ == '__main__':
    main()

